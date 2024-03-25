import re
import openpyxl


with open('reviews-no-blanklines.txt', 'r') as f:
    comments = ['']
    for line in f.readlines():
        if line[0] == '#':
            comments.append(line)
        else:
            comments[-1] += (line)
    print(f'got {len(comments)} comments')


one_worders = []
test_related = []
rest = []


def is_one_word(comment):
    return re.match("^#\d+\s+\w+:\s*\w+[.!?-]*$", comment)


def is_test_related(comment):
    comment_in_lower = comment.lower()
    return 'test' in comment_in_lower or 'assert' in comment_in_lower

keyword_counts = {}

def count_keyword(keyword):
    if keyword in keyword_counts:
        keyword_counts[keyword] += 1
    else:
        keyword_counts[keyword] = 0

def classify(comment):
    classifications = [{
        'class': 'language',
        'keys': ['const', 'value', 'reference', '_ptr', 'point', '++', 'loop',
                 'std', 'return', 'free', 'delete', 'macro', 'class', 'bool',
                 'define', 'compil', 'variable', 'except', 'parameter', 'size',
                 'statement', 'condition', 'block', 'map', 'interface', 'try',
                 'catch', 'final', '==', '||', 'string', 'clause']
    }, {
        'class': 'clean',
        'keys': ['print', 'configur', 'file', 'declar', 'read', 'analysis', 'clean'
                 'duplicat', 'standard', 'comment', 'name', 'repeat', 'reuse', 'comment'
                 'check', 'null', 'spell', 'tab', 'indent', 'naming', 'format', 'log',
                 'dead', 'simplify', 'copyright']
    }, {
        'class': 'system',
        'keys': ['thread', 'memory', 'design', 'requirement', 'need', 'break', 'coef', 
                 'build', 'compute', 'blend', 'locali', 'measur', 'hardware', 'timeout',
                 'state', 'system', 'connect', 'calculat']
    }]
    lower_comment = comment.lower()
    for classifier in classifications:
        for keyword in classifier['keys']:
            if keyword in lower_comment:
                count_keyword(keyword)
                return classifier['class'], keyword
    return 'unclassified', ''


for comment in comments:
    if is_one_word(comment):
        one_worders.append(comment)
    elif is_test_related(comment):
        test_related.append(comment)
    else:
        rest.append(comment)


print(f'{len(one_worders)} one word comments')
print(f'{len(test_related)} related to tests')
print(f'{len(rest)} rest of them')

with open('test_related.txt', 'w') as f:
    f.write('\n'.join(test_related))
with open('unclassified-reviews.txt', 'w') as f:
    f.write('\n'.join(rest))

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.cell(row=1, column=1, value='comment')
sheet.cell(row=1, column=2, value='type')
sheet.cell(row=1, column=3, value='keyword')
for row_index, comment in enumerate(rest):
    row_number = row_index + 2
    print(f'\r{row_index+1}', end='')
    sheet.cell(row=row_number, column=1, value=comment)
    classification, keyword = classify(comment)
    sheet.cell(row=row_number, column=2, value=classification)
    sheet.cell(row=row_number, column=3, value=keyword)
workbook.save('classified-reviews.xlsx')
print(' classified\n')
